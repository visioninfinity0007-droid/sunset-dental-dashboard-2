"""
Export a local dashboard snapshot from an Excel workbook.

Usage:
    python scripts/export_dashboard_seed.py
    python scripts/export_dashboard_seed.py "path/to/workbook.xlsx"

Default source:
    G:\\Coding projects\\Chatbot\\Sunset_Dental_Bot_Database_5.xlsx
"""

import json
import sys
from datetime import UTC, date, datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = Path(r"G:\Coding projects\Chatbot\Sunset_Dental_Bot_Database_5.xlsx")
OUTPUT_PATH = ROOT / "data" / "dashboard-data.json"


def norm(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def clean_name(name):
    raw = str(name or "").strip()
    if not raw or raw.lower() in {"yes please", "unknown", "\u200eunknown"}:
        return "Unknown lead"
    return raw


def sheet_rows(workbook, sheet_name):
    if sheet_name not in workbook.sheetnames:
        print(f"  WARNING: sheet '{sheet_name}' not found, skipping.")
        return []
    ws = workbook[sheet_name]
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [str(h).strip() if h is not None else "" for h in next(rows)]
    except StopIteration:
        return []
    items = []
    for row in rows:
        if not row or not any(c is not None and str(c).strip() for c in row):
            continue
        item = {h: norm(row[i] if i < len(row) else "") for i, h in enumerate(headers) if h}
        items.append(item)
    return items


def resolve_source():
    if len(sys.argv) > 1 and sys.argv[1].strip():
        return Path(sys.argv[1]).expanduser()
    return DEFAULT_SOURCE


def main():
    src = resolve_source()
    print(f"Loading: {src}")
    if not src.exists():
        print(f"ERROR: File not found: {src}")
        sys.exit(1)

    wb = load_workbook(src, read_only=True, data_only=True)
    leads_raw = sheet_rows(wb, "leads")
    chat_raw = sheet_rows(wb, "chat_log")
    appts_raw = sheet_rows(wb, "appointments")

    leads = [
        {
            "phone": str(r.get("phone", "")).strip(),
            "name": clean_name(r.get("name")),
            "stage": str(r.get("stage", "")).strip(),
            "conversationStage": str(r.get("conversation_stage", "")).strip(),
            "intentLevel": str(r.get("intent_level", "")).strip().lower(),
            "leadScore": int(r.get("lead_score") or 0),
            "inquiryType": str(r.get("inquiry_type", "")).strip(),
            "treatmentType": str(r.get("treatment_type", "")).strip(),
            "painLevel": int(r.get("pain_level") or 0),
            "urgencyScore": int(r.get("urgency_score") or 0),
            "preferredDay": str(r.get("preferred_day", "")).strip(),
            "preferredTime": str(r.get("preferred_time", "")).strip(),
            "appointmentTime": str(r.get("appointment_time", "")).strip(),
            "appointmentStatus": str(r.get("appointment_status", "")).strip(),
            "lastIntent": str(r.get("last_intent", "")).strip(),
            "lastMessage": str(r.get("last_message", "")).strip(),
            "lastQuestionAsked": str(r.get("last_question_asked", "")).strip(),
            "lastInteractionAt": str(r.get("last_interaction_at", "")).strip(),
            "sourceChannel": str(r.get("source_channel", "")).strip(),
            "sourceCampaign": str(r.get("source_campaign", "")).strip(),
            "language": str(r.get("language", "")).strip(),
            "currentHandler": str(r.get("current_handler", "")).strip(),
            "handoffStatus": str(r.get("handoff_status", "")).strip(),
            "conversationSummary": str(r.get("conversation_summary", "")).strip(),
            "createdAt": str(r.get("created_at", "")).strip(),
            "updatedAt": str(r.get("updated_at", "")).strip(),
            "appointmentSlotIso": str(r.get("appointment_slot_iso", "")).strip(),
            "noShowCount": int(r.get("no_show_count") or 0),
        }
        for r in leads_raw
    ]

    messages = [
        {
            "timestamp": str(r.get("timestamp", "")).strip(),
            "phone": str(r.get("phone", "")).strip(),
            "direction": str(r.get("direction", "")).strip(),
            "message": str(r.get("message", "")).strip(),
            "intent": str(r.get("intent", "")).strip(),
            "stageAtSend": str(r.get("stage_at_send", "")).strip(),
            "language": str(r.get("language", "")).strip(),
            "mediaType": str(r.get("media_type", "")).strip(),
        }
        for r in chat_raw
    ]

    appointments = [
        {
            "appointmentId": str(r.get("appointment_id", "")).strip(),
            "phone": str(r.get("phone", "")).strip(),
            "name": clean_name(r.get("name")),
            "treatmentType": str(r.get("treatment_type", "")).strip(),
            "slotIso": str(r.get("slot_iso", "")).strip(),
            "slotHuman": str(r.get("slot_human", "")).strip(),
            "status": str(r.get("status", "")).strip(),
            "createdAt": str(r.get("created_at", "")).strip(),
            "reminded24hAt": str(r.get("reminded_24h_at", "")).strip(),
            "reminded2hAt": str(r.get("reminded_2h_at", "")).strip(),
            "outcomeNotes": str(r.get("outcome_notes", "")).strip(),
            "sourceLeadScore": int(r.get("source_lead_score") or 0),
        }
        for r in appts_raw
    ]

    payload = {
        "generatedAt": datetime.now(UTC).isoformat().replace("+00:00", "Z"),
        "meta": {
            "leadCount": len(leads),
            "messageCount": len(messages),
            "appointmentCount": len(appointments),
            "sourceWorkbook": str(src),
        },
        "leads": leads,
        "messages": messages,
        "appointments": appointments,
    }

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"✓ Wrote {OUTPUT_PATH}")
    print(f"  {len(leads)} leads · {len(messages)} messages · {len(appointments)} appointments")


if __name__ == "__main__":
    main()
